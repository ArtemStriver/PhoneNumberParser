#!/usr/bin/env python3
from openpyxl import load_workbook

"""
PhoneNumberParser - модуль, в котором сосредоточен весь функционал программы.
Он производит изъятие данных из файлов формата .txt или .xlsx,
их форматирование, сортировку и запись отформатированных данных по 
разным файлам.
"""


class PhoneNumberParser:

    def __init__(self, path_from, path_out, style_mode, cells):
        self.path_from = path_from
        self.path_out = path_out
        self.style_mode = style_mode
        self.index = ['-', '+', '(', ')', "'", "\n", ' ', ',']
        self.cells = cells #['JF', 'JG', 'JH', 'JI']
        self.numbers = []
        self.correct_numbers = []
        self.incorrect_numbers = []

        self.correct_numbers_kz = []
        self.correct_numbers_belarus = []
        self.correct_numbers_uzbeckistan = []
        self.correct_numbers_kirgistan = []
        self.correct_numbers_tadgikistan = []
        self.correct_numbers_azerbaidgan = []
        self.correct_numbers_italy = []
        self.correct_numbers_germany = []
        self.correct_numbers_turky = []
        self.correct_numbers_izrail = []
        self.correct_numbers_greatbritan = []

        self.correct_number_count = 0

    def read_file_excel(self):
        """Чтение списка номеров из файла формата xlsx."""
        book = load_workbook(filename=self.path_from, data_only=True)
        sheep = book.active
        max_rows = sheep.max_row
        for i in range(2, max_rows + 1):
            for cell in self.cells:  # cell - имя столбца
                number = sheep[cell + str(i)].value  # данные из определенной ячейки
                if not number:
                    continue
                self.numbers.append(number)
        print(len(self.numbers))

    def read_file_txt(self):
        """Чтение списка номеров из файла формата txt."""
        with open(file=self.path_from, mode='r', encoding='utf8') as file:
            for line in file:
                self.numbers.append(line)
        print(len(self.numbers))

    def choose_style(self, number, mode=1):
        """Функция выбора стиля номера."""
        style = ''
        if mode == 1:
            style = '+7' + number
        elif mode == 2:
            style = '+7(' + number[0:3] + ')' + number[3:6] + '-' + number[6:8] + '-' + number[8:]
        return style

    def file_correction(self):
        """Форматирование номеров и сортировка на русские/казахские и иностранные (СНГ и другие)."""
        for odd_numb in self.numbers:
            clear_numb = odd_numb.translate({ord(i): None for i in self.index})
            clear_numb = ''.join(i for i in clear_numb if not i.isalpha())
            if len(clear_numb) == 22:  # два русских номера, но есть варианты номеров русский+иностранный (или наоборот)
                self.correct_number_count += 2
                self.correct_numbers.append(self.choose_style(clear_numb[1:11], self.style_mode) +
                                            '\n' + self.choose_style(clear_numb[12:], self.style_mode))
            elif self.sorting_foreign_numbers(clear_numb):
                pass
            elif len(clear_numb) != 11:
                self.incorrect_numbers.append(clear_numb)
            elif clear_numb[1] != '9' and len(clear_numb) == 11:
                self.correct_numbers_kz.append(self.choose_style(clear_numb[1:], self.style_mode))
            else:
                self.correct_number_count += 1
                self.correct_numbers.append(self.choose_style(clear_numb[1:], self.style_mode))

    def sorting_foreign_numbers(self, numb):
        """Функция сортировки списка иностранных номеров по странам."""
        all_foreign_codes = {'375': self.correct_numbers_belarus,
                             '998': self.correct_numbers_uzbeckistan,
                             '996': self.correct_numbers_kirgistan,
                             '992': self.correct_numbers_tadgikistan,
                             '994': self.correct_numbers_azerbaidgan,
                             '39': self.correct_numbers_italy,
                             '49': self.correct_numbers_germany,
                             '90': self.correct_numbers_turky,
                             '972': self.correct_numbers_izrail,
                             '44': self.correct_numbers_greatbritan}
        for key in all_foreign_codes.keys():
            if numb[0:3] == key:
                all_foreign_codes[key].append(numb)
                return True
            elif numb[0:2] == key:
                all_foreign_codes[key].append(numb)
                return True
        else:
            return False

    def write_file(self):
        """Запись номеров в файлы."""
        all_foreign_numbers = {'Беларусь': self.correct_numbers_belarus,
                               'Узбекистан': self.correct_numbers_uzbeckistan,
                               'Киргизстан': self.correct_numbers_kirgistan,
                               'Таджикистан': self.correct_numbers_tadgikistan,
                               'Азербайджан': self.correct_numbers_azerbaidgan,
                               'Италия': self.correct_numbers_italy,
                               'Германия': self.correct_numbers_germany,
                               'Турция': self.correct_numbers_turky,
                               'Израйль': self.correct_numbers_izrail,
                               'Великобритания': self.correct_numbers_greatbritan}
        with open(file=f'{self.path_out}result_RU_KZ.txt', mode='w', encoding='utf8') as out_file:
            for numb in self.correct_numbers:
                out_file.write(numb + '\n')
            for numb in self.correct_numbers_kz:
                out_file.write(numb + '\n')

        with open(file=f'{self.path_out}result_other.txt', mode='w', encoding='utf8') as out_file:
            for value in all_foreign_numbers.values():
                for numb in value:
                    out_file.write('+' + str(numb) + '\n')

        with open(file=f'{self.path_out}result_other_with_explanations.txt', mode='w', encoding='utf8') as out_file:
            for key, value in all_foreign_numbers.items():
                out_file.write(key + '\n')
                for numb in value:
                    out_file.write(numb + '\n')

        with open(file=f'{self.path_out}incorrect_numbers.txt', mode='w', encoding='utf8') as out_file:
            for numb in self.incorrect_numbers:
                out_file.write('+' + str(numb) + '\n')

    def run(self):
        """Запуск программы."""
        if self.path_from[-4:] == '.txt':
            self.read_file_txt()
        else:
            self.read_file_excel()
        self.file_correction()
        self.write_file()


def run_parsing(path_from, path_out, style_mode, cells):
    """Функция для графического интерфейса."""
    try:
        parser = PhoneNumberParser(path_from=path_from, path_out=path_out, style_mode=style_mode, cells=cells)
        parser.run()
    except Exception as exc:
        print(exc)
